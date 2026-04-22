"""
Student Management Portal
College Campus Management System
Built with FastAPI + SQLAlchemy
"""

from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, BackgroundTasks, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy import create_engine, Column, Integer, String, Text, Float, Date, DateTime, Boolean, func
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime, date, timedelta
import io
import os
import json
import csv
import re
import secrets
import hashlib
import httpx
import math
from openpyxl import load_workbook

# ============== ENVIRONMENT & CONFIG ==============
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
DATABASE_URL = "sqlite:///./campus.db"

DEMO_USERS = [
    {
        "email": "admin@college.edu",
        "password": "admin123",
        "name": "Dr. Admin Kumar",
        "roll_number": "ADMIN001",
        "role": "ADMIN",
        "dept": "ALL",
        "section": "ALL",
        "sem": 0,
    },
    {
        "email": "student1@college.edu",
        "password": "password123",
        "name": "Rahul Sharma",
        "roll_number": "21CSE101",
        "role": "STUDENT",
        "dept": "CSE",
        "section": "A",
        "sem": 4,
    },
    {
        "email": "student2@college.edu",
        "password": "password123",
        "name": "Priya Singh",
        "roll_number": "21CSE102",
        "role": "STUDENT",
        "dept": "CSE",
        "section": "A",
        "sem": 4,
    },
    {
        "email": "student3@college.edu",
        "password": "password123",
        "name": "Amit Patel",
        "roll_number": "21ECE101",
        "role": "STUDENT",
        "dept": "ECE",
        "section": "B",
        "sem": 6,
    },
    {
        "email": "student4@college.edu",
        "password": "password123",
        "name": "Sneha Gupta",
        "roll_number": "23ME101",
        "role": "STUDENT",
        "dept": "ME",
        "section": "A",
        "sem": 2,
    },
    {
        "email": "student5@college.edu",
        "password": "password123",
        "name": "Vikram Reddy",
        "roll_number": "21CSE201",
        "role": "STUDENT",
        "dept": "CSE",
        "section": "B",
        "sem": 4,
    },
]

SAMPLE_SUBJECTS_BY_SCOPE = {
    ("CSE", 4): [
        "Data Structures",
        "DBMS",
        "Operating Systems",
        "Computer Networks",
        "Software Engineering",
    ],
    ("ECE", 6): [
        "VLSI Design",
        "Digital Signal Processing",
        "Embedded Systems",
        "Microcontrollers",
        "Communication Systems",
    ],
    ("ME", 2): [
        "Engineering Mechanics",
        "Thermodynamics",
        "Material Science",
        "Workshop Practice",
        "Engineering Drawing",
    ],
}

DEFAULT_SAMPLE_SUBJECTS = [
    "Mathematics",
    "Programming Fundamentals",
    "Physics",
    "Communication Skills",
    "Environmental Studies",
]

SAMPLE_TIME_SLOTS = [
    "09:00-10:00",
    "10:00-11:00",
    "11:15-12:15",
    "13:00-14:00",
    "14:00-15:00",
]

SAMPLE_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Load College Data from JSON
def load_college_data():
    try:
        with open("college_data.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

COLLEGE_DATA = load_college_data()

MAX_UPLOAD_SIZE_BYTES = 5 * 1024 * 1024
BULK_REPORTS_DIR = "bulk_upload_reports"

# ============== DATABASE SETUP ==============
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ============== MODULAR DATABASE MODELS ==============

class User(Base):
    """User model for students and admins"""
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String(100), unique=True, index=True)
    password = Column(String(255))
    name = Column(String(100))
    username = Column(String(50), unique=True, index=True, nullable=True)
    roll_number = Column(String(20), unique=True, nullable=True)
    role = Column(String(20), default="STUDENT")  # ADMIN or STUDENT
    dept = Column(String(50))
    section = Column(String(10))
    sem = Column(Integer)

class Attendance(Base):
    """Detailed attendance records per class"""
    __tablename__ = "attendance"
    id = Column(Integer, primary_key=True, index=True)
    student_email = Column(String(100), index=True)
    subject_name = Column(String(100))
    date = Column(Date, index=True)
    status = Column(String(10))  # PRESENT, ABSENT, LATE
    dept = Column(String(50))
    section = Column(String(10))
    sem = Column(Integer)

class AttendanceSummary(Base):
    """Aggregated attendance summary per subject"""
    __tablename__ = "attendance_summary"
    id = Column(Integer, primary_key=True, index=True)
    student_email = Column(String(100), index=True)
    subject_name = Column(String(100), index=True)
    attended = Column(Integer, default=0)
    total = Column(Integer, default=0)
    threshold_target = Column(Float, default=75.0)  # Default 75%
    dept = Column(String(50))
    section = Column(String(10))
    sem = Column(Integer)
    last_updated = Column(DateTime, default=datetime.utcnow)

class Timetable(Base):
    """Class timetable entries"""
    __tablename__ = "timetable"
    id = Column(Integer, primary_key=True, index=True)
    day_of_week = Column(String(20))  # Monday, Tuesday, etc.
    period_slots = Column(String(20))  # e.g., "9:00-10:00"
    subject_name = Column(String(100))
    room_number = Column(String(20))
    faculty_name = Column(String(100), nullable=True)
    resource_details = Column(Text, nullable=True)
    dept = Column(String(50))
    section = Column(String(10))
    sem = Column(Integer)

class Announcement(Base):
    """College announcements for students"""
    __tablename__ = "announcements"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(200))
    body = Column(Text)  # Renamed from 'content' as per requirements
    image_url = Column(String(500), nullable=True)
    date = Column(DateTime, default=datetime.utcnow)  # Changed from created_at
    target_dept = Column(String(50))  # "ALL" for all departments
    target_depts = Column(Text, nullable=True)  # CSV format e.g. "CSE,ECE"
    target_sections = Column(Text, nullable=True)  # CSV format e.g. "A,B"
    target_semesters = Column(Text, nullable=True)  # CSV format e.g. "2,4,6"
    priority = Column(String(20), default="normal")  # low, normal, high, urgent


class Resource(Base):
    """Academic resources shared with students"""
    __tablename__ = "resources"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(200))
    description = Column(Text, nullable=True)
    resource_type = Column(String(50), default="DOCUMENT")  # DOCUMENT, LINK, VIDEO, OTHER
    resource_url = Column(String(500), nullable=True)
    dept = Column(String(50), default="ALL")
    section = Column(String(10), default="ALL")
    sem = Column(Integer, default=0)  # 0 means all semesters
    uploaded_by = Column(String(100), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class AcademicUnit(Base):
    """Dynamic department/section combinations managed by admin."""
    __tablename__ = "academic_units"
    id = Column(Integer, primary_key=True, index=True)
    dept = Column(String(50), index=True)
    section = Column(String(10), index=True)
    created_by = Column(String(100), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

def migrate_schema_if_needed():
    """Add newly introduced columns for existing SQLite databases."""
    with engine.connect() as conn:
        timetable_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(timetable)").fetchall()}
        if "resource_details" not in timetable_cols:
            conn.exec_driver_sql("ALTER TABLE timetable ADD COLUMN resource_details TEXT")

        announcement_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(announcements)").fetchall()}
        if "target_depts" not in announcement_cols:
            conn.exec_driver_sql("ALTER TABLE announcements ADD COLUMN target_depts TEXT")
        if "target_sections" not in announcement_cols:
            conn.exec_driver_sql("ALTER TABLE announcements ADD COLUMN target_sections TEXT")
        if "target_semesters" not in announcement_cols:
            conn.exec_driver_sql("ALTER TABLE announcements ADD COLUMN target_semesters TEXT")

        user_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(users)").fetchall()}
        if "username" not in user_cols:
            conn.exec_driver_sql("ALTER TABLE users ADD COLUMN username VARCHAR(50)")


# Create all tables
Base.metadata.create_all(bind=engine)
migrate_schema_if_needed()

# ============== PYDANTIC SCHEMAS ==============

class LoginRequest(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    id: int
    email: str
    name: str
    username: Optional[str] = None
    roll_number: Optional[str] = None
    role: str
    dept: str
    section: str
    sem: int

    class Config:
        from_attributes = True

class AttendanceRecord(BaseModel):
    """Individual subject attendance record"""
    subject_name: str
    attended: int
    total: int
    percentage: float
    threshold_target: float
    status: str  # SAFE, WARNING, DANGER, NO_DATA
    classes_needed: Optional[int] = None
    safe_bunks: Optional[int] = None

class AttendanceResponse(BaseModel):
    """Complete attendance data for a student"""
    overall_percentage: float
    overall_status: str
    subjects: List[AttendanceRecord]
    last_updated: Optional[datetime] = None

class TimetableResponse(BaseModel):
    """Timetable entry response"""
    id: int
    day_of_week: str
    period_slots: str
    subject_name: str
    room_number: str
    faculty_name: Optional[str] = None
    resource_details: Optional[str] = None

    class Config:
        from_attributes = True

class AnnouncementResponse(BaseModel):
    """Announcement card response"""
    id: int
    title: str
    body: str
    image_url: Optional[str] = None
    date: Optional[datetime] = None
    target_dept: str
    target_depts: Optional[List[str]] = None
    target_sections: Optional[List[str]] = None
    target_semesters: Optional[List[int]] = None
    priority: Optional[str] = "normal"

    class Config:
        from_attributes = True

class DashboardResponse(BaseModel):
    """Complete dashboard data"""
    timetable: List[TimetableResponse]
    announcements: List[AnnouncementResponse]
    attendance: Optional[AttendanceResponse] = None

class ChatRequest(BaseModel):
    message: str
    user_email: str

class ChatResponse(BaseModel):
    response: str

class QuickSyncResponse(BaseModel):
    """Quick Sync endpoint response with latest data"""
    attendance: AttendanceResponse
    announcements: List[AnnouncementResponse]
    sync_timestamp: datetime
    data_freshness: str  # "fresh", "updated", "cached"

# Create schemas for admin operations
class AnnouncementCreate(BaseModel):
    title: str
    body: str
    target_dept: str
    target_depts: Optional[List[str]] = None
    target_sections: Optional[List[str]] = None
    target_semesters: Optional[List[int]] = None
    image_url: Optional[str] = None
    priority: Optional[str] = "normal"

class TimetableCreate(BaseModel):
    day_of_week: str
    period_slots: str
    subject_name: str
    room_number: str
    faculty_name: Optional[str] = None
    resource_details: Optional[str] = None
    dept: str
    section: str
    sem: int


class TimetableMatrixCell(BaseModel):
    subject_name: str
    room_number: str
    faculty_name: Optional[str] = None
    resource_details: Optional[str] = None


class TimetableMatrixResponse(BaseModel):
    time_slots: List[str]
    days: List[str]
    matrix: Dict[str, Dict[str, Optional[TimetableMatrixCell]]]


class ResourceCreate(BaseModel):
    title: str
    description: Optional[str] = None
    resource_type: str = "DOCUMENT"
    resource_url: Optional[str] = None
    dept: str = "ALL"
    section: str = "ALL"
    sem: int = 0


class ResourceResponse(BaseModel):
    id: int
    title: str
    description: Optional[str] = None
    resource_type: str
    resource_url: Optional[str] = None
    dept: str
    section: str
    sem: int
    uploaded_by: Optional[str] = None
    created_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class AcademicUnitCreate(BaseModel):
    dept: str
    section: str


class StudentCreateRequest(BaseModel):
    name: str
    email: str
    roll_number: str
    dept: str
    section: str = "A"
    sem: int = 1
    password: Optional[str] = None

# ============== ATTENDANCE CALCULATION ENGINE ==============

def calculate_attendance_percentage(attended: int, total: int) -> float:
    """Calculate percentage attendance"""
    if total == 0:
        return 0.0
    return round((attended / total) * 100, 2)

def get_attendance_status(percentage: float, threshold: float = 75.0) -> str:
    """Determine attendance status based on percentage"""
    if percentage >= threshold:
        return "SAFE"
    elif percentage >= (threshold - 10):
        return "WARNING"
    else:
        return "DANGER"

def calculate_classes_needed_to_reach_target(
    attended: int, 
    total: int, 
    target_percentage: float = 75.0
) -> int:
    """
    Calculate consecutive classes needed to reach target percentage.
    
    Formula: (attended + x) / (total + x) >= target/100
    Solving for x: x >= (target * total - 100 * attended) / (100 - target)
    """
    if total == 0:
        return int(math.ceil(target_percentage / 100))
    
    current_percentage = (attended / total) * 100
    
    # Already at target
    if current_percentage >= target_percentage:
        return 0
    
    numerator = (target_percentage * total) - (100 * attended)
    denominator = 100 - target_percentage
    
    classes_needed = math.ceil(numerator / denominator)
    return max(0, classes_needed)

def calculate_safe_bunks(
    attended: int,
    total: int,
    threshold: float = 75.0
) -> int:
    """
    Calculate how many classes can be missed while maintaining threshold.
    
    Formula: (attended) / (total + x) >= threshold/100
    Solving for x: x <= (100 * attended - threshold * total) / threshold
    """
    if total == 0:
        return 0
    
    current_percentage = (attended / total) * 100
    
    # Below threshold
    if current_percentage < threshold:
        return 0
    
    numerator = (100 * attended) - (threshold * total)
    denominator = threshold
    
    safe_bunks = int(math.floor(numerator / denominator))
    return max(0, safe_bunks)


def csv_to_list(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    return [item.strip() for item in raw.split(",") if item and item.strip()]


def int_csv_to_list(raw: Optional[str]) -> List[int]:
    values = []
    for item in csv_to_list(raw):
        try:
            values.append(int(item))
        except ValueError:
            continue
    return values


def list_to_csv(values: Optional[List[Any]]) -> Optional[str]:
    if not values:
        return None
    cleaned = [str(v).strip() for v in values if str(v).strip()]
    return ",".join(cleaned) if cleaned else None


def get_admin_or_403(db: Session, admin_email: str):
    normalized_email = (admin_email or "").strip().strip('"').strip("'").lower()
    if not normalized_email:
        raise HTTPException(status_code=403, detail="Admin access required")

    admin = db.query(User).filter(
        (func.lower(User.email) == normalized_email) |
        (func.lower(func.coalesce(User.username, "")) == normalized_email)
    ).first()
    if not admin or str(admin.role or "").strip().upper() != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    return admin


def assert_role_barrier(admin: User, dept: str, section: str):
    dept_ok = admin.dept == "ALL" or admin.dept == dept
    section_ok = admin.section == "ALL" or admin.section == section
    if not (dept_ok and section_ok):
        raise HTTPException(status_code=403, detail="Not allowed outside your assigned department/section")


def ensure_demo_users_if_empty(db: Session):
    """Bootstrap demo users for first-run deployments where DB starts empty."""
    if db.query(User).count() > 0:
        return

    users = []
    for user_payload in DEMO_USERS:
        payload = dict(user_payload)
        payload["password"] = hash_password(payload["password"])
        payload["username"] = payload.get("roll_number")
        users.append(User(**payload))

    db.add_all(users)
    db.commit()


def get_sample_subjects_for_scope(dept: str, sem: int) -> List[str]:
    return SAMPLE_SUBJECTS_BY_SCOPE.get((dept, sem), DEFAULT_SAMPLE_SUBJECTS)


def ensure_sample_timetable_for_scope(db: Session, dept: str, section: str, sem: int):
    has_rows = db.query(Timetable).filter(
        Timetable.dept == dept,
        Timetable.section == section,
        Timetable.sem == sem,
    ).first()
    if has_rows:
        return

    subjects = get_sample_subjects_for_scope(dept, sem)
    rows = []
    for day_index, day_name in enumerate(SAMPLE_DAYS):
        for slot_index, slot in enumerate(SAMPLE_TIME_SLOTS):
            subject = subjects[(day_index + slot_index) % len(subjects)]
            rows.append(
                Timetable(
                    day_of_week=day_name,
                    period_slots=slot,
                    subject_name=subject,
                    room_number=f"{dept}-{section}-R{slot_index + 1}",
                    faculty_name=f"Prof. {subject.split()[0]}",
                    resource_details=f"Bring notes for {subject}",
                    dept=dept,
                    section=section,
                    sem=sem,
                )
            )

    db.add_all(rows)
    db.commit()


def ensure_sample_attendance_for_student(db: Session, user: User):
    has_summary = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == user.email
    ).first()
    if has_summary:
        return

    subjects = get_sample_subjects_for_scope(user.dept, user.sem)
    summaries = []
    details = []

    for index, subject in enumerate(subjects):
        total = 28 + (index * 2)
        attended = total - (index + 2)

        summaries.append(
            AttendanceSummary(
                student_email=user.email,
                subject_name=subject,
                attended=attended,
                total=total,
                threshold_target=75.0,
                dept=user.dept,
                section=user.section,
                sem=user.sem,
                last_updated=datetime.utcnow(),
            )
        )

        for offset in range(10):
            class_date = date.today() - timedelta(days=offset)
            status = "PRESENT" if (offset + index) % 5 != 0 else "ABSENT"
            details.append(
                Attendance(
                    student_email=user.email,
                    subject_name=subject,
                    date=class_date,
                    status=status,
                    dept=user.dept,
                    section=user.section,
                    sem=user.sem,
                )
            )

    db.add_all(summaries + details)
    db.commit()


def ensure_sample_resources_for_scope(db: Session, dept: str, section: str, sem: int):
    has_scope_resource = db.query(Resource).filter(
        Resource.dept.in_(["ALL", dept]),
        Resource.section.in_(["ALL", section]),
        Resource.sem.in_([0, sem]),
    ).first()
    if has_scope_resource:
        return

    subjects = get_sample_subjects_for_scope(dept, sem)
    resources = [
        Resource(
            title=f"{subjects[0]} Quick Revision Notes",
            description="Concise notes covering key concepts for exam preparation.",
            resource_type="DOCUMENT",
            resource_url="https://example.com/resources/revision-notes.pdf",
            dept=dept,
            section=section,
            sem=sem,
            uploaded_by="admin@college.edu",
            created_at=datetime.utcnow(),
        ),
        Resource(
            title=f"{subjects[1]} Practice Questions",
            description="Practice set with mixed difficulty and answer key.",
            resource_type="DOCUMENT",
            resource_url="https://example.com/resources/practice-questions.pdf",
            dept=dept,
            section=section,
            sem=sem,
            uploaded_by="admin@college.edu",
            created_at=datetime.utcnow(),
        ),
        Resource(
            title="Campus Learning Portal",
            description="Common resource portal shared across all departments.",
            resource_type="LINK",
            resource_url="https://example.com/portal",
            dept="ALL",
            section="ALL",
            sem=0,
            uploaded_by="admin@college.edu",
            created_at=datetime.utcnow(),
        ),
    ]

    db.add_all(resources)
    db.commit()


def ensure_sample_student_data(db: Session, user: User):
    if user.role != "STUDENT":
        return

    ensure_sample_timetable_for_scope(db, user.dept, user.section, user.sem)
    ensure_sample_attendance_for_student(db, user)
    ensure_sample_resources_for_scope(db, user.dept, user.section, user.sem)


def announcement_visible_for_user(announcement: Announcement, user: User) -> bool:
    user_dept = (user.dept or "").strip().upper()
    user_section = (user.section or "").strip().upper()
    target_dept = (announcement.target_dept or "").strip().upper()

    # Normalize targeting lists to support legacy mixed-case data.
    depts = [(dept or "").strip().upper() for dept in csv_to_list(announcement.target_depts)]
    sections = [(section or "").strip().upper() for section in csv_to_list(announcement.target_sections)]
    semesters = int_csv_to_list(announcement.target_semesters)

    # Legacy rows may have an empty target_dept but still be global announcements.
    if target_dept in ["", "ALL"] and not depts and not sections and not semesters:
        return True

    dept_match = (target_dept in ["", "ALL"]) or (target_dept == user_dept)
    if depts:
        dept_match = user_dept in depts

    section_match = True if not sections else (user_section in sections)
    semester_match = True if not semesters else (user.sem in semesters)

    return dept_match and section_match and semester_match


def serialize_announcement(announcement: Announcement) -> Dict[str, Any]:
    return {
        "id": announcement.id,
        "title": announcement.title,
        "body": announcement.body,
        "image_url": announcement.image_url,
        "date": announcement.date,
        "target_dept": announcement.target_dept,
        "target_depts": csv_to_list(announcement.target_depts),
        "target_sections": csv_to_list(announcement.target_sections),
        "target_semesters": int_csv_to_list(announcement.target_semesters),
        "priority": announcement.priority,
    }


def resource_visible_for_user(resource: Resource, user: User) -> bool:
    dept_match = resource.dept in ["ALL", user.dept]
    section_match = resource.section in ["ALL", user.section]
    sem_match = resource.sem in [0, user.sem]
    return dept_match and section_match and sem_match


def normalize_excel_header(header_value: Any) -> str:
    """Normalize Excel header text into a comparable snake_case-like token."""
    if header_value is None:
        return ""
    header = str(header_value).strip().lower()
    return "".join(ch if ch.isalnum() else "_" for ch in header).strip("_")


def coalesce_header_index(headers: List[str], aliases: List[str]) -> Optional[int]:
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def hash_password(password: str) -> str:
    """Store passwords using salted PBKDF2-HMAC SHA256."""
    iterations = 120000
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def verify_password(plain_password: str, stored_password: str) -> bool:
    """Support secure hashes while keeping backward compatibility with legacy plain-text records."""
    if not stored_password:
        return False

    if not stored_password.startswith("pbkdf2_sha256$"):
        return plain_password == stored_password

    try:
        _, iterations_raw, salt, digest_hex = stored_password.split("$", 3)
        iterations = int(iterations_raw)
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            plain_password.encode("utf-8"),
            salt.encode("utf-8"),
            iterations,
        ).hex()
        return secrets.compare_digest(digest, digest_hex)
    except Exception:
        return False


def generate_default_password(length: int = 10) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$", email or ""))


def normalize_department(raw_department: str) -> str:
    value = (raw_department or "").strip().upper()
    aliases = {
        "COMPUTER SCIENCE": "CSE",
        "COMPUTER SCIENCE ENGINEERING": "CSE",
        "C.S.E": "CSE",
        "ELECTRONICS": "ECE",
        "ELECTRONICS AND COMMUNICATION": "ECE",
        "ELECTRONICS AND COMMUNICATION ENGINEERING": "ECE",
        "E.C.E": "ECE",
        "MECHANICAL": "ME",
        "MECHANICAL ENGINEERING": "ME",
    }
    return aliases.get(value, value)


def normalize_section(raw_section: str) -> str:
    return (raw_section or "").strip().upper()


def parse_semester_value(raw_sem: str) -> int:
    value = (raw_sem or "").strip()
    if not value:
        return 1
    try:
        return int(value)
    except ValueError:
        try:
            parsed = float(value)
            if parsed.is_integer():
                return int(parsed)
        except ValueError:
            pass
    raise ValueError("Semester must be an integer")


def ensure_academic_units_seeded(db: Session):
    """Seed academic_units from existing student/admin data if table is empty."""
    if db.query(AcademicUnit).count() > 0:
        return

    seen = set()
    users = db.query(User).all()
    for user in users:
        dept = normalize_department(user.dept)
        section = normalize_section(user.section)
        if not dept or not section or dept == "ALL" or section == "ALL":
            continue
        key = (dept, section)
        if key in seen:
            continue
        db.add(AcademicUnit(dept=dept, section=section, created_by="system"))
        seen.add(key)

    db.commit()


def get_dynamic_academic_options(db: Session) -> Dict[str, Any]:
    ensure_academic_units_seeded(db)

    depts = set()
    sections = set()
    dept_sections: Dict[str, set] = {}

    for unit in db.query(AcademicUnit).all():
        dept = normalize_department(unit.dept)
        section = normalize_section(unit.section)
        if not dept or not section:
            continue
        depts.add(dept)
        sections.add(section)
        dept_sections.setdefault(dept, set()).add(section)

    # Include values from existing users to avoid stale scope options.
    for user in db.query(User).all():
        dept = normalize_department(user.dept)
        section = normalize_section(user.section)
        if not dept or not section or dept == "ALL" or section == "ALL":
            continue
        depts.add(dept)
        sections.add(section)
        dept_sections.setdefault(dept, set()).add(section)

    return {
        "departments": sorted(depts),
        "sections": sorted(sections),
        "dept_sections": {dept: sorted(values) for dept, values in sorted(dept_sections.items())},
    }

def get_student_attendance(db: Session, email: str, threshold: float = 75.0) -> AttendanceResponse:
    """
    Comprehensive attendance calculation for a student.
    Returns overall and subject-wise attendance with actionable insights.
    """
    summaries = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == email
    ).all()
    
    if not summaries:
        return AttendanceResponse(
            overall_percentage=0,
            overall_status="NO_DATA",
            subjects=[]
        )
    
    subject_records = []
    total_attended = 0
    total_classes = 0
    last_updated = None
    
    for summary in summaries:
        percentage = calculate_attendance_percentage(summary.attended, summary.total)
        status = get_attendance_status(percentage, summary.threshold_target)
        
        classes_needed = None
        safe_bunks = None
        
        if percentage < summary.threshold_target:
            classes_needed = calculate_classes_needed_to_reach_target(
                summary.attended, 
                summary.total, 
                summary.threshold_target
            )
        else:
            safe_bunks = calculate_safe_bunks(
                summary.attended,
                summary.total,
                summary.threshold_target
            )
        
        subject_records.append(AttendanceRecord(
            subject_name=summary.subject_name,
            attended=summary.attended,
            total=summary.total,
            percentage=percentage,
            threshold_target=summary.threshold_target,
            status=status,
            classes_needed=classes_needed,
            safe_bunks=safe_bunks
        ))
        
        total_attended += summary.attended
        total_classes += summary.total
        
        if last_updated is None or summary.last_updated > last_updated:
            last_updated = summary.last_updated
    
    overall_percentage = calculate_attendance_percentage(total_attended, total_classes)
    overall_status = get_attendance_status(overall_percentage, threshold)
    
    return AttendanceResponse(
        overall_percentage=overall_percentage,
        overall_status=overall_status,
        subjects=subject_records,
        last_updated=last_updated
    )

# ============== FASTAPI APP ==============
app = FastAPI(
    title="Student Management Portal",
    version="3.0",
    description="Professional Mobile-First Campus Management System"
)

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============== CREATE DATABASE TABLES ==============
Base.metadata.create_all(bind=engine)

# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ============== AUTHENTICATION ENDPOINTS ==============

@app.post("/login", response_model=UserResponse)
def login(request: LoginRequest, db: Session = Depends(get_db)):
    """Login endpoint - returns user profile"""
    ensure_demo_users_if_empty(db)

    # Normalize email to avoid trivial input mismatch from whitespace/case.
    normalized_email = request.email.strip().lower()
    user = db.query(User).filter(User.email == normalized_email).first()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not verify_password(request.password, user.password):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    ensure_sample_student_data(db, user)
    return user

# ============== DASHBOARD ENDPOINTS ==============

@app.get("/student/dashboard", response_model=DashboardResponse)
def get_student_dashboard(email: str, db: Session = Depends(get_db)):
    """Get personalized dashboard for a student"""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get timetable
    timetable = db.query(Timetable).filter(
        Timetable.dept == user.dept,
        Timetable.section == user.section,
        Timetable.sem == user.sem
    ).all()
    
    # Get announcements
    all_announcements = db.query(Announcement).order_by(Announcement.date.desc()).all()
    announcements = [
        serialize_announcement(ann)
        for ann in all_announcements
        if announcement_visible_for_user(ann, user)
    ]
    
    # Get attendance
    attendance = get_student_attendance(db, email)
    
    return {
        "timetable": timetable,
        "announcements": announcements,
        "attendance": attendance
    }

# ============== ATTENDANCE ENDPOINTS ==============

@app.get("/student/attendance", response_model=AttendanceResponse)
def get_attendance(email: str, db: Session = Depends(get_db)):
    """Get detailed attendance for a student"""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return get_student_attendance(db, email)

@app.get("/student/attendance/{subject_name}")
def get_subject_attendance(email: str, subject_name: str, db: Session = Depends(get_db)):
    """Get attendance for a specific subject"""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    summary = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == email,
        AttendanceSummary.subject_name == subject_name
    ).first()
    
    if not summary:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    percentage = calculate_attendance_percentage(summary.attended, summary.total)
    status = get_attendance_status(percentage, summary.threshold_target)
    classes_needed = calculate_classes_needed_to_reach_target(
        summary.attended, 
        summary.total, 
        summary.threshold_target
    ) if percentage < summary.threshold_target else None
    safe_bunks = calculate_safe_bunks(
        summary.attended,
        summary.total,
        summary.threshold_target
    ) if percentage >= summary.threshold_target else None
    
    return {
        "subject_name": summary.subject_name,
        "attended": summary.attended,
        "total": summary.total,
        "percentage": percentage,
        "threshold_target": summary.threshold_target,
        "status": status,
        "classes_needed": classes_needed,
        "safe_bunks": safe_bunks,
        "message": (
            f"Attend {classes_needed} consecutive classes to reach {summary.threshold_target}%"
            if classes_needed else 
            f"You can miss {safe_bunks} classes and still stay above {summary.threshold_target}%"
        )
    }


def _month_range(reference_date: date) -> tuple[date, date]:
    """Return [start_of_month, start_of_next_month) for date filters."""
    month_start = reference_date.replace(day=1)
    if month_start.month == 12:
        next_month_start = date(month_start.year + 1, 1, 1)
    else:
        next_month_start = date(month_start.year, month_start.month + 1, 1)
    return month_start, next_month_start


def get_monthly_attendance_overview(db: Session, email: str, reference_date: Optional[date] = None) -> Dict[str, Any]:
    """
    Build a month-wise attendance snapshot.
    Uses detailed attendance rows when available; falls back to summary totals.
    """
    today = reference_date or date.today()
    month_start, next_month_start = _month_range(today)

    monthly_rows = db.query(Attendance).filter(
        Attendance.student_email == email,
        Attendance.date >= month_start,
        Attendance.date < next_month_start,
    ).all()

    source = "attendance"
    if monthly_rows:
        total_classes = len(monthly_rows)
        attended_classes = sum(1 for row in monthly_rows if (row.status or "").upper() in {"PRESENT", "LATE"})
    else:
        # Fallback for setups that maintain only summary-level attendance.
        source = "summary_fallback"
        summary_rows = db.query(AttendanceSummary).filter(
            AttendanceSummary.student_email == email
        ).all()
        total_classes = sum(row.total for row in summary_rows)
        attended_classes = sum(row.attended for row in summary_rows)

    percentage = calculate_attendance_percentage(attended_classes, total_classes)
    month_end_day = (next_month_start - timedelta(days=1)).day

    return {
        "month": month_start.strftime("%B %Y"),
        "total_classes": total_classes,
        "attended_classes": attended_classes,
        "percentage": percentage,
        "is_month_end": today.day == month_end_day,
        "source": source,
    }


@app.get("/student/attendance-alerts")
def get_attendance_alerts(email: str, db: Session = Depends(get_db)):
    """Return proactive attendance alerts for chatbot and dashboard notifications."""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    attendance = get_student_attendance(db, email)
    low_subjects = [
        {
            "subject_name": subject.subject_name,
            "percentage": subject.percentage,
            "classes_needed": subject.classes_needed,
        }
        for subject in attendance.subjects
        if subject.percentage < 75
    ]

    return {
        "has_low_attendance": len(low_subjects) > 0,
        "low_attendance_subjects": low_subjects,
        "monthly_summary": get_monthly_attendance_overview(db, email),
    }

# ============== QUICK SYNC ENDPOINT ==============

@app.post("/student/quick-sync", response_model=QuickSyncResponse)
def quick_sync(email: str, db: Session = Depends(get_db)):
    """
    Quick Sync endpoint - Simulates fetching fresh data from ERP like Anveshna.
    Returns latest attendance and announcements in one call.
    """
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get updated attendance
    attendance = get_student_attendance(db, email)
    
    # Get latest announcements
    all_announcements = db.query(Announcement).order_by(Announcement.date.desc()).all()
    announcements = [
        serialize_announcement(ann)
        for ann in all_announcements
        if announcement_visible_for_user(ann, user)
    ][:10]
    
    return {
        "attendance": attendance,
        "announcements": announcements,
        "sync_timestamp": datetime.utcnow(),
        "data_freshness": "fresh"
    }

# ============== TIMETABLE ENDPOINTS ==============

@app.get("/student/timetable")
def get_student_timetable(email: str, db: Session = Depends(get_db)):
    """Get timetable for a student"""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    timetable = db.query(Timetable).filter(
        Timetable.dept == user.dept,
        Timetable.section == user.section,
        Timetable.sem == user.sem
    ).order_by(
        Timetable.day_of_week
    ).all()
    
    return {"timetable": timetable}


@app.get("/student/timetable-matrix", response_model=TimetableMatrixResponse)
def get_student_timetable_matrix(email: str, db: Session = Depends(get_db)):
    """Get timetable as matrix: rows=time slots, columns=days."""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    entries = db.query(Timetable).filter(
        Timetable.dept == user.dept,
        Timetable.section == user.section,
        Timetable.sem == user.sem
    ).all()

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    time_slots = sorted(list({entry.period_slots for entry in entries}))
    matrix: Dict[str, Dict[str, Optional[Dict[str, Any]]]] = {
        slot: {day: None for day in days} for slot in time_slots
    }

    for entry in entries:
        if entry.period_slots in matrix and entry.day_of_week in matrix[entry.period_slots]:
            matrix[entry.period_slots][entry.day_of_week] = {
                "subject_name": entry.subject_name,
                "room_number": entry.room_number,
                "faculty_name": entry.faculty_name,
                "resource_details": entry.resource_details,
            }

    return {
        "time_slots": time_slots,
        "days": days,
        "matrix": matrix,
    }

# ============== ANNOUNCEMENT ENDPOINTS ==============

@app.get("/announcements")
def get_announcements(dept: Optional[str] = None, db: Session = Depends(get_db)):
    """Get announcements"""
    query = db.query(Announcement)
    
    if dept:
        query = query.filter(
            (Announcement.target_dept == dept) | (Announcement.target_dept == "ALL")
        )
    
    return [serialize_announcement(ann) for ann in query.order_by(Announcement.date.desc()).all()]


# ============== RESOURCE ENDPOINTS ==============

@app.get("/student/resources", response_model=List[ResourceResponse])
def get_student_resources(email: str, db: Session = Depends(get_db)):
    """Get resources visible for the student based on dept/section/semester."""
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    resources = db.query(Resource).order_by(Resource.created_at.desc()).all()
    return [resource for resource in resources if resource_visible_for_user(resource, user)]


@app.post("/admin/resources")
def create_resource(data: ResourceCreate, admin_email: str, db: Session = Depends(get_db)):
    """Create a resource (admin only with dept/section barriers)."""
    admin = get_admin_or_403(db, admin_email)

    if data.dept != "ALL" and data.section != "ALL":
        assert_role_barrier(admin, data.dept, data.section)
    else:
        if admin.dept != "ALL" and data.dept not in ["ALL", admin.dept]:
            raise HTTPException(status_code=403, detail="Cannot target other departments")
        if admin.section != "ALL" and data.section not in ["ALL", admin.section]:
            raise HTTPException(status_code=403, detail="Cannot target other sections")

    resource = Resource(
        title=data.title,
        description=data.description,
        resource_type=(data.resource_type or "DOCUMENT").upper(),
        resource_url=data.resource_url,
        dept=data.dept,
        section=data.section,
        sem=data.sem,
        uploaded_by=admin.email,
        created_at=datetime.utcnow(),
    )
    db.add(resource)
    db.commit()
    db.refresh(resource)
    return {"success": True, "message": "Resource created", "id": resource.id}


@app.post("/admin/resources/upload")
async def upload_resource(
    title: str = Form(...),
    description: str = Form(default=""),
    file: UploadFile = File(...),
    admin_email: str = Form(...),
    dept: str = Form(default="ALL"),
    section: str = Form(default="ALL"),
    sem: int = Form(default=0),
    db: Session = Depends(get_db)
):
    """Upload a resource file (PDF, DOCX, etc)"""
    admin = get_admin_or_403(db, admin_email)
    
    # Create uploads directory if it doesn't exist
    os.makedirs("uploads", exist_ok=True)
    
    # Validate file extension
    allowed_extensions = {'.pdf', '.doc', '.docx', '.ppt', '.pptx', '.xls', '.xlsx', '.txt', '.zip'}
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"File type not allowed. Allowed: {allowed_extensions}")
    
    # Save file with unique name
    file_id = secrets.token_hex(8)
    file_path = f"uploads/{file_id}_{file.filename}"
    
    try:
        # Save uploaded file
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
        
        # Create resource record
        resource = Resource(
            title=title,
            description=description,
            resource_type="DOCUMENT",
            resource_url=f"/resources/download/{file_id}",
            dept=dept,
            section=section,
            sem=sem,
            uploaded_by=admin.email,
            created_at=datetime.utcnow(),
        )
        db.add(resource)
        db.commit()
        db.refresh(resource)
        
        return {
            "success": True,
            "message": "Resource uploaded successfully",
            "id": resource.id,
            "download_url": f"/resources/download/{file_id}"
        }
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/resources/download/{file_id}")
def download_resource(file_id: str):
    """Download uploaded resource file"""
    try:
        # Find file in uploads directory
        upload_dir = "uploads"
        if not os.path.exists(upload_dir):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Find file matching the ID
        for filename in os.listdir(upload_dir):
            if filename.startswith(file_id):
                file_path = os.path.join(upload_dir, filename)
                return FileResponse(
                    file_path,
                    filename=filename.split('_', 1)[1],
                    media_type='application/octet-stream'
                )
        
        raise HTTPException(status_code=404, detail="File not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/resources")
def get_all_resources_admin(
    admin_email: str,
    filter_depts: Optional[str] = None,
    filter_sections: Optional[str] = None,
    filter_sems: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """List manageable resources based on admin scope."""
    admin = get_admin_or_403(db, admin_email)
    resources = db.query(Resource).order_by(Resource.created_at.desc()).all()
    scoped_depts = csv_to_list(filter_depts)
    scoped_sections = csv_to_list(filter_sections)
    scoped_sems = int_csv_to_list(filter_sems)

    filtered = []
    for resource in resources:
        if admin.dept != "ALL" and resource.dept not in ["ALL", admin.dept]:
            continue
        if admin.section != "ALL" and resource.section not in ["ALL", admin.section]:
            continue
        if scoped_depts and resource.dept not in ["ALL", *scoped_depts]:
            continue
        if scoped_sections and resource.section not in ["ALL", *scoped_sections]:
            continue
        if scoped_sems and resource.sem not in [0, *scoped_sems]:
            continue
        filtered.append(resource)

    return {"resources": filtered, "count": len(filtered)}


@app.get("/admin/resources/{resource_id}")
def get_resource_detail(resource_id: int, admin_email: str, db: Session = Depends(get_db)):
    admin = get_admin_or_403(db, admin_email)
    resource = db.query(Resource).filter(Resource.id == resource_id).first()
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")
    if admin.dept != "ALL" and resource.dept not in ["ALL", admin.dept]:
        raise HTTPException(status_code=403, detail="Cannot access this resource")
    if admin.section != "ALL" and resource.section not in ["ALL", admin.section]:
        raise HTTPException(status_code=403, detail="Cannot access this resource")
    return resource


@app.put("/admin/resources/{resource_id}")
def update_resource(resource_id: int, data: ResourceCreate, admin_email: str, db: Session = Depends(get_db)):
    admin = get_admin_or_403(db, admin_email)
    resource = db.query(Resource).filter(Resource.id == resource_id).first()
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")

    if admin.dept != "ALL" and resource.dept not in ["ALL", admin.dept]:
        raise HTTPException(status_code=403, detail="Cannot update this resource")
    if admin.section != "ALL" and resource.section not in ["ALL", admin.section]:
        raise HTTPException(status_code=403, detail="Cannot update this resource")

    if data.dept != "ALL" and data.section != "ALL":
        assert_role_barrier(admin, data.dept, data.section)

    resource.title = data.title
    resource.description = data.description
    resource.resource_type = (data.resource_type or "DOCUMENT").upper()
    resource.resource_url = data.resource_url
    resource.dept = data.dept
    resource.section = data.section
    resource.sem = data.sem

    db.commit()
    db.refresh(resource)
    return {"success": True, "message": "Resource updated", "id": resource.id}


@app.delete("/admin/resources/{resource_id}")
def delete_resource(resource_id: int, admin_email: str, db: Session = Depends(get_db)):
    admin = get_admin_or_403(db, admin_email)
    resource = db.query(Resource).filter(Resource.id == resource_id).first()
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")

    if admin.dept != "ALL" and resource.dept not in ["ALL", admin.dept]:
        raise HTTPException(status_code=403, detail="Cannot delete this resource")
    if admin.section != "ALL" and resource.section not in ["ALL", admin.section]:
        raise HTTPException(status_code=403, detail="Cannot delete this resource")

    db.delete(resource)
    db.commit()
    return {"success": True, "message": "Resource deleted"}

# ============== ADMIN ENDPOINTS ==============

# ============== TIMETABLE MANAGEMENT (CREATE, READ, UPDATE, DELETE) ==============

@app.post("/admin/timetable", response_model=dict)
def create_timetable_entry(data: TimetableCreate, admin_email: str, db: Session = Depends(get_db)):
    """Create a new timetable entry"""
    admin = get_admin_or_403(db, admin_email)
    assert_role_barrier(admin, data.dept, data.section)

    entry = Timetable(
        day_of_week=data.day_of_week,
        period_slots=data.period_slots,
        subject_name=data.subject_name,
        room_number=data.room_number,
        faculty_name=data.faculty_name,
        resource_details=data.resource_details,
        dept=data.dept,
        section=data.section,
        sem=data.sem
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"success": True, "message": "Timetable entry added", "id": entry.id}

@app.get("/admin/timetable")
def get_all_timetable_entries(
    admin_email: str,
    filter_depts: Optional[str] = None,
    filter_sections: Optional[str] = None,
    filter_sems: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get all timetable entries for admin management"""
    admin = get_admin_or_403(db, admin_email)
    query = db.query(Timetable)
    if admin.dept != "ALL":
        query = query.filter(Timetable.dept == admin.dept)
    if admin.section != "ALL":
        query = query.filter(Timetable.section == admin.section)

    scoped_depts = csv_to_list(filter_depts)
    scoped_sections = csv_to_list(filter_sections)
    scoped_sems = int_csv_to_list(filter_sems)
    if scoped_depts:
        query = query.filter(Timetable.dept.in_(scoped_depts))
    if scoped_sections:
        query = query.filter(Timetable.section.in_(scoped_sections))
    if scoped_sems:
        query = query.filter(Timetable.sem.in_(scoped_sems))

    entries = query.all()
    return {"entries": entries, "count": len(entries)}

@app.get("/admin/timetable/{entry_id}")
def get_timetable_entry(entry_id: int, admin_email: str, db: Session = Depends(get_db)):
    """Get a specific timetable entry"""
    admin = get_admin_or_403(db, admin_email)
    entry = db.query(Timetable).filter(Timetable.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Timetable entry not found")
    assert_role_barrier(admin, entry.dept, entry.section)
    return entry

@app.put("/admin/timetable/{entry_id}")
def update_timetable_entry(entry_id: int, data: TimetableCreate, admin_email: str, db: Session = Depends(get_db)):
    """Update a timetable entry"""
    admin = get_admin_or_403(db, admin_email)
    entry = db.query(Timetable).filter(Timetable.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Timetable entry not found")

    assert_role_barrier(admin, entry.dept, entry.section)
    assert_role_barrier(admin, data.dept, data.section)
    
    entry.day_of_week = data.day_of_week
    entry.period_slots = data.period_slots
    entry.subject_name = data.subject_name
    entry.room_number = data.room_number
    entry.faculty_name = data.faculty_name
    entry.resource_details = data.resource_details
    entry.dept = data.dept
    entry.section = data.section
    entry.sem = data.sem
    
    db.commit()
    db.refresh(entry)
    return {"success": True, "message": "Timetable entry updated", "id": entry.id}

@app.delete("/admin/timetable/{entry_id}")
def delete_timetable_entry(entry_id: int, admin_email: str, db: Session = Depends(get_db)):
    """Delete a timetable entry"""
    admin = get_admin_or_403(db, admin_email)
    entry = db.query(Timetable).filter(Timetable.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Timetable entry not found")

    assert_role_barrier(admin, entry.dept, entry.section)
    
    subject = entry.subject_name
    db.delete(entry)
    db.commit()
    return {"success": True, "message": f"Timetable entry for {subject} deleted"}

# ============== ANNOUNCEMENT MANAGEMENT (CREATE, READ, UPDATE, DELETE) ==============

@app.post("/admin/announcement", response_model=dict)
def create_announcement(data: AnnouncementCreate, admin_email: str, db: Session = Depends(get_db)):
    """Create a new announcement"""
    admin = get_admin_or_403(db, admin_email)
    dept_targets = data.target_depts or []
    section_targets = data.target_sections or []

    if admin.dept != "ALL" and not dept_targets and data.target_dept == "ALL":
        dept_targets = [admin.dept]
    if admin.section != "ALL" and not section_targets:
        section_targets = [admin.section]

    if admin.dept != "ALL" and dept_targets and any(d != admin.dept for d in dept_targets):
        raise HTTPException(status_code=403, detail="Cannot target other departments")
    if admin.section != "ALL" and section_targets and any(s != admin.section for s in section_targets):
        raise HTTPException(status_code=403, detail="Cannot target other sections")

    announcement = Announcement(
        title=data.title,
        body=data.body,
        target_dept=(data.target_dept if data.target_dept != "ALL" else (admin.dept if dept_targets else "ALL")),
        target_depts=list_to_csv(dept_targets),
        target_sections=list_to_csv(section_targets),
        target_semesters=list_to_csv(data.target_semesters),
        image_url=data.image_url,
        priority=data.priority,
        date=datetime.utcnow()
    )
    db.add(announcement)
    db.commit()
    db.refresh(announcement)
    return {"success": True, "message": "Announcement posted", "id": announcement.id}

@app.get("/admin/announcements")
def get_all_announcements_admin(
    admin_email: str,
    filter_depts: Optional[str] = None,
    filter_sections: Optional[str] = None,
    filter_sems: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get all announcements for admin management"""
    admin = get_admin_or_403(db, admin_email)
    announcements = db.query(Announcement).order_by(Announcement.date.desc()).all()
    scoped_depts = csv_to_list(filter_depts)
    scoped_sections = csv_to_list(filter_sections)
    scoped_sems = int_csv_to_list(filter_sems)
    filtered = []
    for ann in announcements:
        dept_targets = csv_to_list(ann.target_depts)
        section_targets = csv_to_list(ann.target_sections)
        sem_targets = int_csv_to_list(ann.target_semesters)
        if admin.dept != "ALL" and not dept_targets and ann.target_dept not in ["ALL", admin.dept]:
            continue
        if admin.dept != "ALL" and dept_targets and admin.dept not in dept_targets:
            continue
        if admin.section != "ALL" and section_targets and admin.section not in section_targets:
            continue

        if scoped_depts:
            ann_depts = dept_targets or ([ann.target_dept] if ann.target_dept != "ALL" else [])
            if ann_depts and not any(dept in scoped_depts for dept in ann_depts):
                continue
        if scoped_sections and section_targets and not any(section in scoped_sections for section in section_targets):
            continue
        if scoped_sems and sem_targets and not any(sem in scoped_sems for sem in sem_targets):
            continue

        filtered.append(serialize_announcement(ann))
    return {"announcements": filtered, "count": len(filtered)}

@app.get("/admin/announcement/{announcement_id}")
def get_announcement_detail(announcement_id: int, admin_email: str, db: Session = Depends(get_db)):
    """Get a specific announcement"""
    get_admin_or_403(db, admin_email)
    announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not announcement:
        raise HTTPException(status_code=404, detail="Announcement not found")
    return serialize_announcement(announcement)

@app.put("/admin/announcement/{announcement_id}")
def update_announcement(announcement_id: int, data: AnnouncementCreate, admin_email: str, db: Session = Depends(get_db)):
    """Update an announcement"""
    admin = get_admin_or_403(db, admin_email)
    announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not announcement:
        raise HTTPException(status_code=404, detail="Announcement not found")

    dept_targets = data.target_depts or []
    section_targets = data.target_sections or []
    if admin.dept != "ALL" and not dept_targets and data.target_dept == "ALL":
        dept_targets = [admin.dept]
    if admin.section != "ALL" and not section_targets:
        section_targets = [admin.section]
    if admin.dept != "ALL" and dept_targets and any(d != admin.dept for d in dept_targets):
        raise HTTPException(status_code=403, detail="Cannot target other departments")
    if admin.section != "ALL" and section_targets and any(s != admin.section for s in section_targets):
        raise HTTPException(status_code=403, detail="Cannot target other sections")
    
    announcement.title = data.title
    announcement.body = data.body
    announcement.target_dept = (data.target_dept if data.target_dept != "ALL" else (admin.dept if dept_targets else "ALL"))
    announcement.target_depts = list_to_csv(dept_targets)
    announcement.target_sections = list_to_csv(section_targets)
    announcement.target_semesters = list_to_csv(data.target_semesters)
    announcement.image_url = data.image_url
    announcement.priority = data.priority
    announcement.date = datetime.utcnow()
    
    db.commit()
    db.refresh(announcement)
    return {"success": True, "message": "Announcement updated", "id": announcement.id}

@app.delete("/admin/announcement/{announcement_id}")
def delete_announcement(announcement_id: int, admin_email: str, db: Session = Depends(get_db)):
    """Delete an announcement"""
    get_admin_or_403(db, admin_email)
    announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not announcement:
        raise HTTPException(status_code=404, detail="Announcement not found")
    
    title = announcement.title
    db.delete(announcement)
    db.commit()
    return {"success": True, "message": f"Announcement '{title}' deleted"}

# ============== ATTENDANCE MANAGEMENT (CREATE, READ, UPDATE, DELETE) ==============

@app.post("/admin/attendance", response_model=dict)
def record_attendance(
    admin_email: str,
    student_email: str,
    subject_name: str,
    attended: int,
    total: int,
    db: Session = Depends(get_db)
):
    """Record or create new attendance for a student"""
    admin = get_admin_or_403(db, admin_email)
    user = db.query(User).filter(User.email == student_email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    assert_role_barrier(admin, user.dept, user.section)
    
    summary = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == student_email,
        AttendanceSummary.subject_name == subject_name
    ).first()
    
    if summary:
        summary.attended = attended
        summary.total = total
        summary.last_updated = datetime.utcnow()
    else:
        summary = AttendanceSummary(
            student_email=student_email,
            subject_name=subject_name,
            attended=attended,
            total=total,
            dept=user.dept,
            section=user.section,
            sem=user.sem,
            last_updated=datetime.utcnow()
        )
        db.add(summary)
    
    db.commit()
    db.refresh(summary)
    return {"success": True, "message": "Attendance recorded"}

@app.get("/admin/attendance")
def get_all_attendance(
    admin_email: str,
    filter_depts: Optional[str] = None,
    filter_sections: Optional[str] = None,
    filter_sems: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get all attendance records for admin management"""
    admin = get_admin_or_403(db, admin_email)
    query = db.query(AttendanceSummary)
    if admin.dept != "ALL":
        query = query.filter(AttendanceSummary.dept == admin.dept)
    if admin.section != "ALL":
        query = query.filter(AttendanceSummary.section == admin.section)

    scoped_depts = csv_to_list(filter_depts)
    scoped_sections = csv_to_list(filter_sections)
    scoped_sems = int_csv_to_list(filter_sems)
    if scoped_depts:
        query = query.filter(AttendanceSummary.dept.in_(scoped_depts))
    if scoped_sections:
        query = query.filter(AttendanceSummary.section.in_(scoped_sections))
    if scoped_sems:
        query = query.filter(AttendanceSummary.sem.in_(scoped_sems))

    records = query.order_by(
        AttendanceSummary.student_email, 
        AttendanceSummary.subject_name
    ).all()
    return {"records": records, "count": len(records)}

@app.get("/admin/attendance/{student_email}")
def get_student_attendance_records(student_email: str, admin_email: str, db: Session = Depends(get_db)):
    """Get all attendance records for a specific student"""
    admin = get_admin_or_403(db, admin_email)
    records = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == student_email
    ).all()
    if not records:
        raise HTTPException(status_code=404, detail="No attendance records found for this student")

    assert_role_barrier(admin, records[0].dept, records[0].section)
    return {"records": records, "count": len(records)}

@app.put("/admin/attendance/{student_email}/{subject_name}")
def update_attendance(
    student_email: str,
    subject_name: str,
    admin_email: str,
    attended: int,
    total: int,
    db: Session = Depends(get_db)
):
    """Update attendance for a student in a specific subject"""
    admin = get_admin_or_403(db, admin_email)
    user = db.query(User).filter(User.email == student_email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    assert_role_barrier(admin, user.dept, user.section)
    
    summary = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == student_email,
        AttendanceSummary.subject_name == subject_name
    ).first()
    
    if not summary:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    summary.attended = attended
    summary.total = total
    summary.last_updated = datetime.utcnow()
    
    db.commit()
    db.refresh(summary)
    return {"success": True, "message": f"Attendance for {subject_name} updated"}

@app.delete("/admin/attendance/{student_email}/{subject_name}")
def delete_attendance_record(student_email: str, subject_name: str, admin_email: str, db: Session = Depends(get_db)):
    """Delete an attendance record"""
    admin = get_admin_or_403(db, admin_email)
    record = db.query(AttendanceSummary).filter(
        AttendanceSummary.student_email == student_email,
        AttendanceSummary.subject_name == subject_name
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="Attendance record not found")

    assert_role_barrier(admin, record.dept, record.section)
    
    db.delete(record)
    db.commit()
    return {"success": True, "message": f"Attendance record for {subject_name} deleted"}


@app.post("/admin/attendance/upload-excel")
async def upload_attendance_excel(
    admin_email_form: Optional[str] = Form(None, alias="admin_email"),
    admin_email_query: Optional[str] = Query(None, alias="admin_email"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload attendance summary using Excel headers:
    student_email, subject_name, attended, total, dept, section, sem
    """
    admin_email = admin_email_form or admin_email_query or ""
    admin = get_admin_or_403(db, admin_email)
    workbook = load_workbook(filename=io.BytesIO(await file.read()), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel is empty")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required = ["student_email", "subject_name", "attended", "total", "dept", "section", "sem"]
    if not all(name in headers for name in required):
        raise HTTPException(status_code=400, detail=f"Missing headers. Required: {', '.join(required)}")

    idx = {name: headers.index(name) for name in required}
    updated = 0
    created = 0

    for row in rows[1:]:
        if not row or row[idx["student_email"]] is None:
            continue

        student_email = str(row[idx["student_email"]]).strip()
        subject_name = str(row[idx["subject_name"]]).strip()
        attended = int(row[idx["attended"]] or 0)
        total = int(row[idx["total"]] or 0)
        dept = str(row[idx["dept"]]).strip()
        section = str(row[idx["section"]]).strip()
        sem = int(row[idx["sem"]] or 0)

        assert_role_barrier(admin, dept, section)

        existing = db.query(AttendanceSummary).filter(
            AttendanceSummary.student_email == student_email,
            AttendanceSummary.subject_name == subject_name
        ).first()

        if existing:
            existing.attended = attended
            existing.total = total
            existing.dept = dept
            existing.section = section
            existing.sem = sem
            existing.last_updated = datetime.utcnow()
            updated += 1
        else:
            db.add(AttendanceSummary(
                student_email=student_email,
                subject_name=subject_name,
                attended=attended,
                total=total,
                dept=dept,
                section=section,
                sem=sem,
                last_updated=datetime.utcnow()
            ))
            created += 1

    db.commit()
    return {"success": True, "created": created, "updated": updated, "message": "Attendance Excel processed"}


@app.post("/admin/timetable/upload-excel")
async def upload_timetable_excel(
    admin_email_form: Optional[str] = Form(None, alias="admin_email"),
    admin_email_query: Optional[str] = Query(None, alias="admin_email"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Supported formats:
    1) Row format: day_of_week, period_slots, subject_name, room_number, faculty_name, resource_details, dept, section, sem
    2) Matrix format: first col "time", day columns (monday..saturday), plus optional dept, section, sem.
       Each cell value: Subject|Room|Faculty|Resource (last three optional).
    """
    admin_email = admin_email_form or admin_email_query or ""
    admin = get_admin_or_403(db, admin_email)
    workbook = load_workbook(filename=io.BytesIO(await file.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel is empty")

    headers = [normalize_excel_header(cell) for cell in rows[0]]
    created = 0

    day_idx = coalesce_header_index(headers, ["day_of_week", "day", "weekday"])
    time_idx = coalesce_header_index(headers, ["period_slots", "period_slot", "time", "time_slot", "timeslot"])

    if day_idx is not None and time_idx is not None:
        subject_idx = coalesce_header_index(headers, ["subject_name", "subject", "course", "course_name"])
        room_idx = coalesce_header_index(headers, ["room_number", "room", "classroom", "location"])
        dept_idx = coalesce_header_index(headers, ["dept", "department"])
        section_idx = coalesce_header_index(headers, ["section"])
        sem_idx = coalesce_header_index(headers, ["sem", "semester"])
        faculty_idx = coalesce_header_index(headers, ["faculty_name", "faculty", "teacher", "staff_name"])
        resource_idx = coalesce_header_index(headers, ["resource_details", "resource", "resources", "notes"])

        required_names = []
        if subject_idx is None:
            required_names.append("subject_name|subject")
        if room_idx is None:
            required_names.append("room_number|room")
        if dept_idx is None:
            required_names.append("dept|department")
        if section_idx is None:
            required_names.append("section")
        if sem_idx is None:
            required_names.append("sem|semester")
        if required_names:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Missing headers for row format. Required aliases include: "
                    + ", ".join(required_names)
                )
            )

        for row in rows[1:]:
            if not row or row[day_idx] is None:
                continue
            dept = str(row[dept_idx]).strip()
            section = str(row[section_idx]).strip()
            sem = int(row[sem_idx] or 0)
            assert_role_barrier(admin, dept, section)

            db.add(Timetable(
                day_of_week=str(row[day_idx]).strip(),
                period_slots=str(row[time_idx]).strip(),
                subject_name=str(row[subject_idx]).strip(),
                room_number=str(row[room_idx]).strip(),
                faculty_name=str(row[faculty_idx]).strip() if faculty_idx is not None and row[faculty_idx] else None,
                resource_details=str(row[resource_idx]).strip() if resource_idx is not None and row[resource_idx] else None,
                dept=dept,
                section=section,
                sem=sem
            ))
            created += 1
    else:
        day_map = {
            "monday": "Monday", "mon": "Monday",
            "tuesday": "Tuesday", "tue": "Tuesday", "tues": "Tuesday",
            "wednesday": "Wednesday", "wed": "Wednesday",
            "thursday": "Thursday", "thu": "Thursday", "thurs": "Thursday",
            "friday": "Friday", "fri": "Friday",
            "saturday": "Saturday", "sat": "Saturday"
        }

        time_idx = coalesce_header_index(headers, ["time", "period_slots", "period_slot", "time_slot", "timeslot"])
        if time_idx is None:
            raise HTTPException(status_code=400, detail="Matrix format requires header: time or period_slots")

        dept_idx = coalesce_header_index(headers, ["dept", "department"])
        section_idx = coalesce_header_index(headers, ["section"])
        sem_idx = coalesce_header_index(headers, ["sem", "semester"])
        day_indices = {header: headers.index(header) for header in headers if header in day_map}

        if not day_indices:
            raise HTTPException(status_code=400, detail="Matrix format requires day columns: monday..saturday (or mon..sat)")

        for row in rows[1:]:
            if not row or row[time_idx] is None:
                continue
            period_slots = str(row[time_idx]).strip()
            dept = str(row[dept_idx]).strip() if dept_idx is not None and row[dept_idx] else admin.dept
            section = str(row[section_idx]).strip() if section_idx is not None and row[section_idx] else admin.section
            sem = int(row[sem_idx] or 0) if sem_idx is not None else 0

            assert_role_barrier(admin, dept, section)

            for day_key, col_idx in day_indices.items():
                cell = row[col_idx]
                if cell is None or str(cell).strip() == "":
                    continue

                parts = [part.strip() for part in str(cell).split("|")]
                subject_name = parts[0] if len(parts) > 0 else ""
                room_number = parts[1] if len(parts) > 1 and parts[1] else "TBD"
                faculty_name = parts[2] if len(parts) > 2 and parts[2] else None
                resource_details = parts[3] if len(parts) > 3 and parts[3] else None

                db.add(Timetable(
                    day_of_week=day_map[day_key],
                    period_slots=period_slots,
                    subject_name=subject_name,
                    room_number=room_number,
                    faculty_name=faculty_name,
                    resource_details=resource_details,
                    dept=dept,
                    section=section,
                    sem=sem
                ))
                created += 1

    db.commit()
    return {"success": True, "created": created, "message": "Timetable Excel processed"}


@app.post("/admin/announcements/upload-excel")
async def upload_announcements_excel(
    admin_email_form: Optional[str] = Form(None, alias="admin_email"),
    admin_email_query: Optional[str] = Query(None, alias="admin_email"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload announcements using Excel headers:
    title, body, priority, image_url, target_dept, target_depts, target_sections, target_semesters
    target_* list headers are comma-separated values.
    """
    admin_email = admin_email_form or admin_email_query or ""
    admin = get_admin_or_403(db, admin_email)
    workbook = load_workbook(filename=io.BytesIO(await file.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel is empty")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required = ["title", "body"]
    if not all(name in headers for name in required):
        raise HTTPException(status_code=400, detail="Required headers: title, body")

    idx = {name: headers.index(name) for name in headers if name}
    created = 0
    for row in rows[1:]:
        if not row or not row[idx["title"]]:
            continue

        target_dept = str(row[idx.get("target_dept", -1)]).strip() if idx.get("target_dept") is not None and row[idx.get("target_dept", 0)] else "ALL"
        target_depts = csv_to_list(str(row[idx.get("target_depts", -1)])) if idx.get("target_depts") is not None and row[idx.get("target_depts", 0)] else []
        target_sections = csv_to_list(str(row[idx.get("target_sections", -1)])) if idx.get("target_sections") is not None and row[idx.get("target_sections", 0)] else []
        target_semesters = int_csv_to_list(str(row[idx.get("target_semesters", -1)])) if idx.get("target_semesters") is not None and row[idx.get("target_semesters", 0)] else []

        if admin.dept != "ALL" and target_depts and any(d != admin.dept for d in target_depts):
            raise HTTPException(status_code=403, detail="Cannot target other departments")
        if admin.section != "ALL" and target_sections and any(s != admin.section for s in target_sections):
            raise HTTPException(status_code=403, detail="Cannot target other sections")

        db.add(Announcement(
            title=str(row[idx["title"]]).strip(),
            body=str(row[idx["body"]]).strip(),
            priority=str(row[idx.get("priority", -1)]).strip() if idx.get("priority") is not None and row[idx.get("priority", 0)] else "normal",
            image_url=str(row[idx.get("image_url", -1)]).strip() if idx.get("image_url") is not None and row[idx.get("image_url", 0)] else None,
            target_dept=target_dept,
            target_depts=list_to_csv(target_depts),
            target_sections=list_to_csv(target_sections),
            target_semesters=list_to_csv(target_semesters),
            date=datetime.utcnow()
        ))
        created += 1

    db.commit()
    return {"success": True, "created": created, "message": "Announcements Excel processed"}


@app.post("/admin/students/bulk-upload")
async def bulk_upload_students(
    admin_email: str = Form(...),
    send_notifications: bool = Form(False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Create many student accounts from CSV/XLSX with row-level validation and error reporting."""
    admin = get_admin_or_403(db, admin_email)

    file_name = (file.filename or "").strip()
    ext = os.path.splitext(file_name)[1].lower()
    if ext not in [".csv", ".xlsx"]:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are allowed")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(file_bytes) > MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=400, detail="File too large. Maximum allowed size is 5 MB")

    rows: List[List[Any]] = []
    headers: List[str] = []

    if ext == ".csv":
        try:
            decoded = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="CSV must be UTF-8 encoded")

        reader = csv.reader(io.StringIO(decoded))
        csv_rows = list(reader)
        if not csv_rows:
            raise HTTPException(status_code=400, detail="CSV is empty")
        headers = [normalize_excel_header(cell) for cell in csv_rows[0]]
        rows = [list(r) for r in csv_rows[1:]]
    else:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        xlsx_rows = list(sheet.iter_rows(values_only=True))
        if not xlsx_rows:
            raise HTTPException(status_code=400, detail="Excel is empty")
        headers = [normalize_excel_header(cell) for cell in xlsx_rows[0]]
        rows = [list(r) for r in xlsx_rows[1:]]

    name_idx = coalesce_header_index(headers, ["name", "full_name", "student_name"])
    email_idx = coalesce_header_index(headers, ["email", "email_address"])
    roll_idx = coalesce_header_index(headers, ["roll_number", "roll", "roll_no", "rollno", "register_number"])
    dept_idx = coalesce_header_index(headers, ["department", "dept"])
    section_idx = coalesce_header_index(headers, ["section"])
    sem_idx = coalesce_header_index(headers, ["sem", "semester"])

    missing = []
    if name_idx is None:
        missing.append("Name")
    if email_idx is None:
        missing.append("Email")
    if roll_idx is None:
        missing.append("Roll Number")
    if dept_idx is None:
        missing.append("Department")
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    existing_students = db.query(User).filter(User.role == "STUDENT").all()
    existing_emails = {str(u.email or "").strip().lower() for u in existing_students if u.email}
    existing_rolls = {str(u.roll_number or "").strip().upper() for u in existing_students if u.roll_number}
    existing_usernames = {str(u.username or "").strip().upper() for u in existing_students if u.username}

    seen_emails: set = set()
    seen_rolls: set = set()
    seen_usernames: set = set()

    created = 0
    errors: List[Dict[str, Any]] = []
    credentials: List[Dict[str, str]] = []

    for row_index, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def get_cell(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        name = get_cell(name_idx)
        email = get_cell(email_idx).lower()
        roll_number = get_cell(roll_idx).upper()
        department = normalize_department(get_cell(dept_idx))
        section = get_cell(section_idx).upper() if section_idx is not None else "A"
        sem_raw = get_cell(sem_idx) if sem_idx is not None else "1"

        if not name or not email or not roll_number or not department:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Missing required fields",
            })
            continue

        if not is_valid_email(email):
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Invalid email format",
            })
            continue

        try:
            sem = parse_semester_value(sem_raw)
        except ValueError:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Semester must be an integer",
            })
            continue

        username = roll_number

        if admin.dept != "ALL" and department != admin.dept:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Department out of admin scope",
            })
            continue

        if admin.section != "ALL" and (section or "A") != admin.section:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Section out of admin scope",
            })
            continue

        if email in existing_emails or email in seen_emails:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Duplicate email",
            })
            continue

        if roll_number in existing_rolls or roll_number in seen_rolls:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Duplicate roll number",
            })
            continue

        if username in existing_usernames or username in seen_usernames:
            errors.append({
                "row_number": row_index,
                "name": name,
                "email": email,
                "roll_number": roll_number,
                "department": department,
                "error": "Duplicate username",
            })
            continue

        default_password = generate_default_password()
        db.add(User(
            name=name,
            email=email,
            username=username,
            roll_number=roll_number,
            dept=department,
            section=section or "A",
            sem=sem if sem > 0 else 1,
            role="STUDENT",
            password=hash_password(default_password),
        ))

        seen_emails.add(email)
        seen_rolls.add(roll_number)
        seen_usernames.add(username)
        created += 1
        credentials.append({
            "name": name,
            "email": email,
            "username": username,
            "roll_number": roll_number,
            "password": default_password,
        })

    db.commit()

    report_id = None
    if errors:
        os.makedirs(BULK_REPORTS_DIR, exist_ok=True)
        report_id = secrets.token_hex(8)
        report_path = os.path.join(BULK_REPORTS_DIR, f"bulk_student_upload_errors_{report_id}.csv")
        with open(report_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(
                csvfile,
                fieldnames=["row_number", "name", "email", "roll_number", "department", "error"],
            )
            writer.writeheader()
            writer.writerows(errors)

    notifications_sent = created if send_notifications else 0
    return {
        "success": len(errors) == 0,
        "message": "Bulk upload completed" if not errors else "Bulk upload completed with row errors",
        "created_count": created,
        "failed_count": len(errors),
        "notifications_requested": send_notifications,
        "notifications_sent": notifications_sent,
        "credentials": credentials,
        "error_report_id": report_id,
        "error_report_download_url": (
            f"/admin/students/bulk-upload/error-report/{report_id}?admin_email={admin_email}"
            if report_id else None
        ),
    }


@app.get("/admin/students/bulk-upload/error-report/{report_id}")
def download_bulk_upload_error_report(report_id: str, admin_email: str, db: Session = Depends(get_db)):
    """Download CSV error report generated during bulk student upload."""
    get_admin_or_403(db, admin_email)

    safe_report_id = re.sub(r"[^a-f0-9]", "", report_id.lower())
    report_path = os.path.join(BULK_REPORTS_DIR, f"bulk_student_upload_errors_{safe_report_id}.csv")
    if not os.path.exists(report_path):
        raise HTTPException(status_code=404, detail="Error report not found")

    return FileResponse(
        report_path,
        media_type="text/csv",
        filename=f"bulk_student_upload_errors_{safe_report_id}.csv",
    )

@app.get("/admin/stats")
def get_admin_stats(db: Session = Depends(get_db)):
    """Get admin dashboard statistics"""
    students = db.query(User).filter(User.role == "STUDENT").count()
    timetable_count = db.query(Timetable).count()
    announcement_count = db.query(Announcement).count()
    
    return {
        "students": students,
        "timetable_entries": timetable_count,
        "announcements": announcement_count
    }


@app.get("/admin/academic-options")
def get_admin_academic_options(admin_email: str, db: Session = Depends(get_db)):
    """Return dynamic department and section options for admin scope/forms."""
    admin = get_admin_or_403(db, admin_email)
    options = get_dynamic_academic_options(db)

    departments = options["departments"]
    sections = options["sections"]
    dept_sections = options["dept_sections"]

    if admin.dept != "ALL":
        departments = [d for d in departments if d == admin.dept]
        dept_sections = {k: v for k, v in dept_sections.items() if k == admin.dept}
    if admin.section != "ALL":
        sections = [s for s in sections if s == admin.section]
        dept_sections = {k: [s for s in values if s == admin.section] for k, values in dept_sections.items()}

    return {
        "departments": departments,
        "sections": sections,
        "dept_sections": dept_sections,
    }


@app.post("/admin/academic-options")
def create_admin_academic_option(data: AcademicUnitCreate, admin_email: str, db: Session = Depends(get_db)):
    """Create new department/section option and make it available immediately in admin scope."""
    admin = get_admin_or_403(db, admin_email)
    dept = normalize_department(data.dept)
    section = normalize_section(data.section)

    if not dept or not section:
        raise HTTPException(status_code=400, detail="Department and section are required")

    if admin.dept != "ALL" and dept != admin.dept:
        raise HTTPException(status_code=403, detail="Cannot create department outside admin scope")
    if admin.section != "ALL" and section != admin.section:
        raise HTTPException(status_code=403, detail="Cannot create section outside admin scope")

    exists = db.query(AcademicUnit).filter(
        AcademicUnit.dept == dept,
        AcademicUnit.section == section,
    ).first()
    if not exists:
        db.add(AcademicUnit(dept=dept, section=section, created_by=admin.email))
        db.commit()

    return {
        "success": True,
        "message": "Academic option saved",
        "dept": dept,
        "section": section,
    }


@app.get("/admin/students")
def list_admin_students(
    admin_email: str,
    dept: Optional[str] = None,
    section: Optional[str] = None,
    sem: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """List students with optional filters for admin student-management tab."""
    admin = get_admin_or_403(db, admin_email)

    query = db.query(User).filter(User.role == "STUDENT")

    if admin.dept != "ALL":
        query = query.filter(User.dept == admin.dept)
    if admin.section != "ALL":
        query = query.filter(User.section == admin.section)

    normalized_dept = normalize_department(dept) if dept else None
    normalized_section = normalize_section(section) if section else None

    if normalized_dept:
        query = query.filter(User.dept == normalized_dept)
    if normalized_section:
        query = query.filter(User.section == normalized_section)
    if sem is not None:
        query = query.filter(User.sem == sem)

    students = query.order_by(User.dept, User.section, User.sem, User.roll_number).all()
    return {
        "students": [
            {
                "id": s.id,
                "name": s.name,
                "email": s.email,
                "roll_number": s.roll_number,
                "dept": s.dept,
                "section": s.section,
                "sem": s.sem,
            }
            for s in students
        ]
    }


@app.post("/admin/students")
def create_admin_student(data: StudentCreateRequest, admin_email: str, db: Session = Depends(get_db)):
    """Create single student account from admin student-management tab."""
    admin = get_admin_or_403(db, admin_email)

    email = (data.email or "").strip().lower()
    roll_number = (data.roll_number or "").strip().upper()
    name = (data.name or "").strip()
    dept = normalize_department(data.dept)
    section = normalize_section(data.section)

    if not name or not email or not roll_number or not dept or not section:
        raise HTTPException(status_code=400, detail="Name, email, roll number, dept and section are required")
    if not is_valid_email(email):
        raise HTTPException(status_code=400, detail="Invalid email format")
    if data.sem < 1:
        raise HTTPException(status_code=400, detail="Semester must be >= 1")

    if admin.dept != "ALL" and dept != admin.dept:
        raise HTTPException(status_code=403, detail="Cannot create student outside admin department")
    if admin.section != "ALL" and section != admin.section:
        raise HTTPException(status_code=403, detail="Cannot create student outside admin section")

    if db.query(User).filter(User.email == email).first():
        raise HTTPException(status_code=400, detail="Email already exists")
    if db.query(User).filter(User.roll_number == roll_number).first():
        raise HTTPException(status_code=400, detail="Roll number already exists")
    if db.query(User).filter(User.username == roll_number).first():
        raise HTTPException(status_code=400, detail="Username conflict for roll number")

    plain_password = (data.password or "").strip() or generate_default_password()
    user = User(
        name=name,
        email=email,
        username=roll_number,
        roll_number=roll_number,
        dept=dept,
        section=section,
        sem=data.sem,
        role="STUDENT",
        password=hash_password(plain_password),
    )
    db.add(user)

    existing_unit = db.query(AcademicUnit).filter(
        AcademicUnit.dept == dept,
        AcademicUnit.section == section,
    ).first()
    if not existing_unit:
        db.add(AcademicUnit(dept=dept, section=section, created_by=admin.email))

    db.commit()

    return {
        "success": True,
        "message": "Student created",
        "credentials": {
            "email": email,
            "username": roll_number,
            "password": plain_password,
        },
    }


@app.put("/admin/students/{student_email}")
def update_admin_student(student_email: str, data: StudentCreateRequest, admin_email: str, db: Session = Depends(get_db)):
    """Update a student account from admin student-management tab."""
    admin = get_admin_or_403(db, admin_email)
    email = (student_email or "").strip().lower()

    student = db.query(User).filter(User.email == email, User.role == "STUDENT").first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if admin.dept != "ALL" and student.dept != admin.dept:
        raise HTTPException(status_code=403, detail="Cannot edit student outside admin department")
    if admin.section != "ALL" and student.section != admin.section:
        raise HTTPException(status_code=403, detail="Cannot edit student outside admin section")

    # Normalize and validate updates
    new_dept = normalize_department(data.dept)
    new_section = normalize_section(data.section)
    new_name = (data.name or "").strip()
    new_email = (data.email or "").strip().lower()
    new_roll = (data.roll_number or "").strip().upper()

    if not new_name or not new_email or not new_roll or not new_dept or not new_section:
        raise HTTPException(status_code=400, detail="All fields are required")
    if not is_valid_email(new_email):
        raise HTTPException(status_code=400, detail="Invalid email format")
    if data.sem < 1:
        raise HTTPException(status_code=400, detail="Semester must be >= 1")

    # Check scope for new department/section
    if admin.dept != "ALL" and new_dept != admin.dept:
        raise HTTPException(status_code=403, detail="Cannot assign student outside admin department")
    if admin.section != "ALL" and new_section != admin.section:
        raise HTTPException(status_code=403, detail="Cannot assign student outside admin section")

    # Verify no duplicate email (if email changed)
    if new_email != email and db.query(User).filter(User.email == new_email).first():
        raise HTTPException(status_code=400, detail="Email already exists")
    # Verify no duplicate roll (if roll changed)
    if new_roll != student.roll_number and db.query(User).filter(User.roll_number == new_roll).first():
        raise HTTPException(status_code=400, detail="Roll number already exists")

    # Update student
    student.name = new_name
    student.email = new_email
    student.roll_number = new_roll
    student.username = new_roll
    student.dept = new_dept
    student.section = new_section
    student.sem = data.sem

    # Auto-create academic unit if needed
    existing_unit = db.query(AcademicUnit).filter(
        AcademicUnit.dept == new_dept,
        AcademicUnit.section == new_section,
    ).first()
    if not existing_unit:
        db.add(AcademicUnit(dept=new_dept, section=new_section, created_by=admin.email))

    db.commit()
    return {
        "success": True,
        "message": "Student updated",
        "student": {
            "id": student.id,
            "name": student.name,
            "email": student.email,
            "roll_number": student.roll_number,
            "dept": student.dept,
            "section": student.section,
            "sem": student.sem,
        }
    }


@app.delete("/admin/students/{student_email}")
def delete_admin_student(student_email: str, admin_email: str, db: Session = Depends(get_db)):
    """Delete a student account from admin student-management tab."""
    admin = get_admin_or_403(db, admin_email)
    email = (student_email or "").strip().lower()

    student = db.query(User).filter(User.email == email, User.role == "STUDENT").first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if admin.dept != "ALL" and student.dept != admin.dept:
        raise HTTPException(status_code=403, detail="Cannot delete student outside admin department")
    if admin.section != "ALL" and student.section != admin.section:
        raise HTTPException(status_code=403, detail="Cannot delete student outside admin section")

    db.delete(student)
    db.commit()
    return {"success": True, "message": "Student deleted"}


@app.put("/admin/academic-options")
def update_admin_academic_option(data: AcademicUnitCreate, old_dept: str, old_section: str, admin_email: str, db: Session = Depends(get_db)):
    """Update an existing department/section combination."""
    admin = get_admin_or_403(db, admin_email)
    
    old_dept = normalize_department(old_dept)
    old_section = normalize_section(old_section)
    new_dept = normalize_department(data.dept)
    new_section = normalize_section(data.section)

    if not old_dept or not old_section or not new_dept or not new_section:
        raise HTTPException(status_code=400, detail="Department and section are required")

    if admin.dept != "ALL" and (old_dept != admin.dept or new_dept != admin.dept):
        raise HTTPException(status_code=403, detail="Cannot modify department outside admin scope")
    if admin.section != "ALL" and (old_section != admin.section or new_section != admin.section):
        raise HTTPException(status_code=403, detail="Cannot modify section outside admin scope")

    # Find the academic unit to update
    unit = db.query(AcademicUnit).filter(
        AcademicUnit.dept == old_dept,
        AcademicUnit.section == old_section,
    ).first()
    
    if not unit:
        raise HTTPException(status_code=404, detail="Department/section combination not found")

    # Check if new combination already exists
    existing = db.query(AcademicUnit).filter(
        AcademicUnit.dept == new_dept,
        AcademicUnit.section == new_section,
    ).first()
    if existing and existing.id != unit.id:
        raise HTTPException(status_code=400, detail="This department/section combination already exists")

    # Update the unit
    unit.dept = new_dept
    unit.section = new_section
    db.commit()

    return {
        "success": True,
        "message": "Department/section updated",
        "dept": new_dept,
        "section": new_section,
    }


@app.delete("/admin/academic-options")
def delete_admin_academic_option(dept: str, section: str, admin_email: str, db: Session = Depends(get_db)):
    """Delete a department/section combination."""
    admin = get_admin_or_403(db, admin_email)
    
    dept = normalize_department(dept)
    section = normalize_section(section)

    if not dept or not section:
        raise HTTPException(status_code=400, detail="Department and section are required")

    if admin.dept != "ALL" and dept != admin.dept:
        raise HTTPException(status_code=403, detail="Cannot delete department outside admin scope")
    if admin.section != "ALL" and section != admin.section:
        raise HTTPException(status_code=403, detail="Cannot delete section outside admin scope")

    unit = db.query(AcademicUnit).filter(
        AcademicUnit.dept == dept,
        AcademicUnit.section == section,
    ).first()
    
    if not unit:
        raise HTTPException(status_code=404, detail="Department/section combination not found")

    # Check if any students are in this dept/section
    students = db.query(User).filter(User.dept == dept, User.section == section, User.role == "STUDENT").count()
    if students > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {students} student(s) assigned to this department/section")

    db.delete(unit)
    db.commit()

    return {"success": True, "message": "Department/section deleted"}

# ============== INTELLIGENT ASSISTANT ==============

ASSISTANT_SYSTEM_PROMPT = """
You are an intelligent Student Assistant at the college.

**Your Persona:**
- Professional, helpful, and empathetic
- Data-driven and precise with student records
- Proactive in identifying attendance shortages
- Encouraging and supportive in tone
- Always action-oriented with clear next steps
- Understand natural language flexibly (short forms, typos, mixed phrasing, Hinglish-style wording)

**Your Expertise:**
1. **Attendance Management**: Use the student's database records to provide precise answers about attendance percentage, classes needed to reach 75%, and attendance buffer
2. **Schedule Planning**: Help students understand their timetable and today's classes
3. **College Information**: Share college policies, facilities, placements details
4. **Announcements**: Keep students informed about important college updates
5. **Academic Guidance**: Provide study tips and resource recommendations

**Core Capabilities:**
- Calculate attendance insights: "You need to attend 5 consecutive classes to reach 75%"
- Show current class schedule: "You have 3 classes today, starting at 9:00 AM"
- Alert about announcements: "There's an important announcement about exam schedules"
- Provide actionable guidance: "Focus on Physics (currently 68%)"

**Guidelines:**
- Always verify data from the student's actual database records
- Be precise with numbers and percentages
- Suggest specific action items
- Use emojis sparingly but effectively (📚, ✅, ⚠️)
- Never require students to ask in a fixed pattern; infer intent from natural language
- If data is unavailable, say so and guide them to take action
- Keep responses concise (2-3 sentences max)
- Always end with a helpful suggestion or question

**Examples of Good Responses:**
- "Your attendance is at 72% - attend the next 4 Physics classes to reach 75%. Safe attendance! 📚"
- "You have 4 classes today: DSA (9:00), OOP (11:00), Database (1:00), Web Dev (3:00). Ready to crush it? ✅"
- "Check out the new placement announcement - TCS is hiring for internships! 🎯"
"""

@app.post("/chat", response_model=ChatResponse)
async def chat_endpoint(request: ChatRequest, db: Session = Depends(get_db)):
    """
    Intelligent Student Assistant
    Uses student data from database for precise, context-aware responses.
    """
    email = (request.user_email or "").strip().lower()
    message = (request.message or "").strip()

    if not email:
        return ChatResponse(response="Please login again so I can identify your student profile.")

    if not message:
        return ChatResponse(response="Please type a question. I can help with attendance, timetable, announcements, and campus info.")

    user = db.query(User).filter(User.email == email).first()
    if not user:
        return ChatResponse(response="User not found. Please login first.")

    if user.role != "STUDENT":
        return ChatResponse(response="This assistant is available only on the student dashboard. Please login with a student account.")
    
    # Gather student context from database
    attendance = get_student_attendance(db, email)
    timetable = db.query(Timetable).filter(
        Timetable.dept == user.dept,
        Timetable.section == user.section,
        Timetable.sem == user.sem
    ).all()
    
    all_announcements = db.query(Announcement).order_by(Announcement.date.desc()).all()
    announcements = [
        ann for ann in all_announcements
        if announcement_visible_for_user(ann, user)
    ][:5]
    monthly_summary = get_monthly_attendance_overview(db, email)
    
    # Build context message
    context_msg = f"""
User: {user.name} ({user.roll_number})
Department: {user.dept}, Section {user.section}, Semester {user.sem}

Current Attendance: {attendance.overall_percentage}% ({attendance.overall_status})
Monthly Attendance Snapshot ({monthly_summary['month']}): {monthly_summary['attended_classes']}/{monthly_summary['total_classes']} ({monthly_summary['percentage']}%)
Low attendance subjects:
"""
    
    for subject in attendance.subjects:
        if subject.percentage < 75:
            context_msg += f"\n- {subject.subject_name}: {subject.percentage}% (need {subject.classes_needed} classes)"

    if announcements:
        context_msg += "\nRecent announcements:"
        for ann in announcements[:3]:
            context_msg += f"\n- {ann.title} ({ann.priority})"
    
    # Use Groq API if available, otherwise generate local response
    if GROQ_API_KEY:
        response_text = await get_assistant_response_from_groq(
            request.message, 
            context_msg, 
            user, 
            attendance, 
            timetable
        )
    else:
        response_text = generate_assistant_local_response(
            request.message,
            user,
            attendance,
            timetable,
            announcements
        )
    
    return ChatResponse(response=response_text)

async def get_assistant_response_from_groq(message: str, context: str, user, attendance, timetable) -> str:
    """Get response from Groq API"""
    try:
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}"}
        
        prompt = f"""{ASSISTANT_SYSTEM_PROMPT}

{context}

Student Question: {message}

Respond with accurate, actionable guidance."""
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json={
                    "model": "mixtral-8x7b-32768",
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 200,
                    "temperature": 0.7
                },
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                return data["choices"][0]["message"]["content"]
    except Exception as e:
        print(f"Groq API error: {e}")
    
    # Fallback to local response
    return generate_assistant_local_response(message, user, attendance, timetable, [])

def generate_assistant_local_response(message: str, user, attendance, timetable, announcements) -> str:
    """Generate intelligent local response based on student data and natural-language intent."""
    msg_lower = message.lower()
    cleaned_message = re.sub(r"[^a-z0-9\s]", " ", msg_lower)
    tokens = {token for token in cleaned_message.split() if token}

    subject_by_name = {s.subject_name.lower(): s for s in attendance.subjects}

    def has_intent(*keywords: str) -> bool:
        """Check intent with phrase-aware and token-aware matching."""
        return any((keyword in msg_lower) or (keyword in tokens) for keyword in keywords)

    def extract_subject_query():
        for subject in attendance.subjects:
            if subject.subject_name.lower() in msg_lower:
                return subject
        return None

    def college_info_response() -> str:
        info = COLLEGE_DATA.get("college_info", {})
        placements = COLLEGE_DATA.get("placements", {}).get("overview", {})
        facilities = COLLEGE_DATA.get("facilities", {})
        dept_data = COLLEGE_DATA.get("departments", {}).get(user.dept, {})

        if has_intent("placement", "placements", "package", "recruiter", "job"):
            if placements:
                return (
                    f"🎯 Placements snapshot: {placements.get('placement_rate', 'N/A')} placement rate, "
                    f"highest package {placements.get('highest_package', 'N/A')}, average {placements.get('average_package', 'N/A')}. "
                    f"Want top recruiters list too?"
                )
            return "🎯 Placement data is not available right now. Check with the placement cell for latest drives."

        if has_intent("library", "hostel", "sports", "canteen", "facility", "facilities", "transport"):
            library_timing = facilities.get("library", {}).get("timing", "N/A")
            hostel_capacity = facilities.get("hostel", {}).get("boys_hostel", {}).get("capacity", 0) + facilities.get("hostel", {}).get("girls_hostel", {}).get("capacity", 0)
            return (
                f"🏫 Campus facilities: Library timing is {library_timing}, hostel capacity is about {hostel_capacity}, "
                f"and sports + transport facilities are available. Need details for a specific facility?"
            )

        if has_intent("department", "hod", "lab", "labs", "specialization", "specializations") and dept_data:
            labs = dept_data.get("labs", [])[:3]
            return (
                f"🏛️ {user.dept} department: HOD is {dept_data.get('hod', 'N/A')}. "
                f"Popular labs include {', '.join(labs) if labs else 'N/A'}. "
                f"Want semester-wise subjects too?"
            )

        if info:
            return (
                f"🏫 {info.get('name', 'Your college')} is in {info.get('location', 'N/A')} and is "
                f"accredited as {info.get('accreditation', 'N/A')}. "
                f"If you want, I can share contacts, fees, facilities, or placement details."
            )

        return "🏫 I can help with college details like placements, facilities, rules, and contacts. Ask me what you need."
    
    if has_intent("hi", "hello", "hey", "good morning", "good afternoon", "good evening"):
        return f"Hi {user.name}! 👋 I'm your student assistant. Ask me about attendance, timetable, announcements, placements, or facilities."

    if has_intent("my name", "who am i", "profile", "about me", "which department", "my department"):
        return (
            f"You are {user.name} ({user.roll_number}), {user.dept}-{user.section}, semester {user.sem}. "
            f"Current overall attendance is {attendance.overall_percentage}%. Need a subject-wise breakdown?"
        )

    # Attendance queries
    if has_intent("attendance", "percentage", "bunk", "safe", "shortage", "low attendance", "drop", "dropped"):
        subject_match = extract_subject_query()
        if subject_match:
            if subject_match.percentage < 75:
                return (
                    f"📉 In {subject_match.subject_name}, your attendance is {subject_match.percentage}% "
                    f"({subject_match.attended}/{subject_match.total}). Attend {subject_match.classes_needed} more classes to reach 75%."
                )
            return (
                f"✅ In {subject_match.subject_name}, your attendance is {subject_match.percentage}% "
                f"({subject_match.attended}/{subject_match.total}). You can safely miss about {subject_match.safe_bunks or 0} class(es)."
            )

        if attendance.overall_status == "NO_DATA":
            return "📊 No attendance data available yet. Once your classes are recorded, I can help you track progress!"
        
        low_subjects = [s for s in attendance.subjects if s.percentage < 75]
        if low_subjects:
            worst = min(low_subjects, key=lambda x: x.percentage)
            return f"⚠️ Your {worst.subject_name} is at {worst.percentage}% - attend {worst.classes_needed} more classes to reach 75%! You've got this! 💪"
        else:
            safe_bunks = sum(s.safe_bunks or 0 for s in attendance.subjects) // len(attendance.subjects) if attendance.subjects else 0
            return f"✅ Great attendance! You're at {attendance.overall_percentage}% overall. You have an attendance buffer of ~{safe_bunks} classes per subject. Keep it up! 🎯"
    
    # Schedule queries
    if has_intent("schedule", "timetable", "class", "classes", "today", "tomorrow", "slot", "period"):
        from datetime import datetime
        today = datetime.now().strftime("%A")
        today_classes = [t for t in timetable if t.day_of_week == today]
        
        if not timetable:
            return "📅 No timetable data found. Ask your admin to upload the schedule!"
        
        if today_classes:
            classes_info = "\n".join([f"• {c.subject_name} at {c.period_slots} in {c.room_number}" for c in today_classes])
            return f"📚 Your classes today ({today}):\n{classes_info}\n\nReady to learn? Let's go! 🚀"
        else:
            return f"🎉 No classes today ({today})! Perfect time for self-study or revision. 📖"

    # Announcements queries
    if has_intent("announcement", "announcements", "notice", "notices", "update", "updates", "news"):
        if not announcements:
            return "📢 No new announcements right now. I can still help with schedule, attendance, or college info."
        latest = announcements[0]
        return f"📢 Latest for {user.dept}-{user.section}: {latest.title}. Want a quick summary of all recent updates?"

    # College information queries
    if has_intent(
        "college", "clg", "campus", "placement", "facilities", "library", "hostel",
        "rules", "regulations", "fees", "department", "hod", "syllabus", "contacts"
    ):
        return college_info_response()

    # Notification/reminder queries
    if has_intent("notify", "notification", "remind", "reminder", "alert", "alerts"):
        low_subjects = [s for s in attendance.subjects if s.percentage < 75]
        if low_subjects:
            worst = min(low_subjects, key=lambda x: x.percentage)
            return f"🔔 Alert set mentally: {worst.subject_name} is currently {worst.percentage}%. Attend {worst.classes_needed} classes to reach 75%."
        return f"🔔 You're currently at {attendance.overall_percentage}% overall, above the 75% threshold. I'll keep highlighting any risk subjects."
    
    # General help
    if has_intent("help", "what can", "how", "support", "assist"):
        return "I'm your smart campus assistant! 🎓 I can help with attendance planning, schedule, announcements, and college info (placements/facilities/rules). What do you need?"
    
    # Default response
    return "I can help with attendance, schedule, announcements, and any college-related questions in normal language. Try asking in your own words, like 'how is my attendance now?' or 'placement details please'."

# ============== INITIALIZATION ==============

@app.get("/")
def root():
    """Serve frontend entry point on root path."""
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    return {
        "service": "Student Management Portal",
        "version": "3.0",
        "status": "running",
        "assistant": "Intelligent Student Assistant"
    }


@app.get("/tailwind.min.css")
def serve_tailwind_css():
    """Serve locally hosted Tailwind stylesheet."""
    if os.path.exists("tailwind.min.css"):
        return FileResponse("tailwind.min.css", media_type="text/css")
    raise HTTPException(status_code=404, detail="tailwind.min.css not found")


@app.get("/script.js")
def serve_script():
    """Serve frontend JavaScript bundle."""
    if os.path.exists("script.js"):
        return FileResponse("script.js")
    raise HTTPException(status_code=404, detail="script.js not found")


@app.get("/index.html")
def serve_index_file():
    """Serve frontend HTML explicitly."""
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    raise HTTPException(status_code=404, detail="index.html not found")
